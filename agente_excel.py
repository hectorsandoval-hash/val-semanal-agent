"""
AGENTE EXCEL - Procesamiento de archivos Excel de valorizacion
==============================================================
Porta la logica de extraccion del index.html (SheetJS) a Python (openpyxl).
Lee las hojas RES-COSTO, RVAL y CURVA del Excel para extraer los datos
necesarios para generar el reporte.
"""
import io
from datetime import datetime
from openpyxl import load_workbook


def procesar(excel_bytes):
    """
    Procesa un archivo Excel de valorizacion semanal.

    Args:
        excel_bytes: bytes del archivo Excel

    Returns:
        dict con toda la data extraida, o None si falla
    """
    try:
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)

        # Validar hojas OBLIGATORIAS (RES-COSTO y RVAL)
        required_sheets = ["RES-COSTO", "RVAL"]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f'No se encontro la pestana "{sheet_name}" en el archivo.')

        data = {
            "resCosto": _extract_res_costo(wb["RES-COSTO"]),
            "rval": _extract_rval(wb["RVAL"]),
        }

        # CURVA es OPCIONAL - si no existe, el reporte se genera sin Curva S
        if "CURVA" in wb.sheetnames:
            data["curva"] = _extract_curva(wb["CURVA"])
            print("  [EXCEL] Hoja CURVA encontrada.")
        else:
            data["curva"] = None
            print(f"  [EXCEL] Hoja CURVA no encontrada (hojas: {wb.sheetnames[:5]}...). Reporte sin Curva S.")

        # Derivar info del proyecto
        data["projectName"] = (
            data["resCosto"]["projectName"]
            or data["rval"]["projectName"]
            or "PROYECTO"
        )
        data["shortName"] = _get_short_name(data["projectName"])
        data["date"] = (
            data["resCosto"]["date"]
            or data["rval"]["date"]
            or datetime.now()
        )
        data["author"] = (
            data["resCosto"]["author"]
            or data["rval"]["author"]
            or ""
        )

        wb.close()
        return data

    except Exception as e:
        print(f"  [EXCEL] Error procesando: {e}")
        raise


# ============================================================
# FUNCIONES AUXILIARES (equivalentes a cellVal/cellNum del JS)
# ============================================================

def _cell_val(ws, addr):
    """Lee el valor de una celda. Equivalente a cellVal() del JS."""
    cell = ws[addr]
    if cell.value is not None:
        return cell.value
    return None


def _cell_num(ws, addr):
    """Lee el valor numerico de una celda. Equivalente a cellNum() del JS."""
    cell = ws[addr]
    if cell.value is None:
        return 0
    val = cell.value
    if isinstance(val, (int, float)):
        return float(val)
    # Intentar parsear string
    try:
        cleaned = str(val).replace(",", "")
        return float(cleaned)
    except (ValueError, TypeError):
        return 0


def _col_letter(col_index):
    """Convierte indice de columna (0-based) a letra. 0->A, 1->B, etc."""
    result = ""
    col = col_index
    while True:
        result = chr(65 + col % 26) + result
        col = col // 26 - 1
        if col < 0:
            break
    return result


# ============================================================
# EXTRACT RES-COSTO
# Porta extractResCosto() del index.html
# ============================================================

def _extract_res_costo(ws):
    """Extrae datos de la hoja RES-COSTO."""
    result = {
        "projectName": "",
        "date": None,
        "author": "",
        "personalObrero": 0,
        "materiales": 0,
        "alquileres": 0,
        "subcontratos": 0,
        "costosVarios": 0,
        "planillaStaff": 0,
        "otrosGG": 0,
    }

    # Project info - scan rows 2-8
    for r in range(2, 9):
        b = _cell_val(ws, f"B{r}")
        c = _cell_val(ws, f"C{r}")
        g = _cell_val(ws, f"G{r}")

        if b and isinstance(b, str) and "Proyecto" in b:
            result["projectName"] = str(c) if c else ""
        if b and isinstance(b, str) and "Elaborado" in b:
            result["author"] = str(c) if c else ""

        # Buscar fecha
        if isinstance(g, datetime):
            result["date"] = g
        else:
            f_val = _cell_val(ws, f"F{r}")
            if f_val and isinstance(f_val, str) and "fecha" in f_val.lower() and g:
                if isinstance(g, datetime):
                    result["date"] = g
                elif isinstance(g, (int, float)) and 40000 < g < 60000:
                    # Excel serial date
                    from datetime import timedelta
                    result["date"] = datetime(1899, 12, 30) + timedelta(days=int(g))

    # Parse cost categories
    category_map = {
        "PERSONAL DE OBRERO": "personalObrero",
        "MATERIALES": "materiales",
        "ALQUILERES": "alquileres",
        "SUBCONTRATO": "subcontratos",
        "COSTOS VARIOS": "costosVarios",
        "COSTO DE OBRA GG": "gg",
    }

    current_category = ""
    max_row = ws.max_row or 100

    for r in range(10, max_row + 1):
        b_val = _cell_val(ws, f"B{r}")
        c_val = _cell_val(ws, f"C{r}")
        d_val = _cell_num(ws, f"D{r}")

        if b_val and isinstance(b_val, str):
            # Intentar parsear como numero
            try:
                float(b_val)
                is_number = True
            except (ValueError, TypeError):
                is_number = False

            if not is_number:
                # Category header row
                upper = b_val.upper()
                for key, field in category_map.items():
                    if key in upper:
                        current_category = field
                        break
                continue

        # Data row with item number
        if b_val is not None and d_val != 0:
            # Verificar si b_val es un numero (item number)
            is_item = False
            if isinstance(b_val, (int, float)):
                is_item = True
            elif isinstance(b_val, str):
                try:
                    float(b_val)
                    is_item = True
                except (ValueError, TypeError):
                    pass

            if is_item:
                if current_category == "gg":
                    desc = (str(c_val) if c_val else "").lower()
                    if "staff" in desc or "planilla staff" in desc:
                        result["planillaStaff"] += d_val
                    else:
                        result["otrosGG"] += d_val
                elif current_category:
                    result[current_category] += d_val

    result["totalCD"] = (
        result["personalObrero"]
        + result["materiales"]
        + result["alquileres"]
        + result["subcontratos"]
        + result["costosVarios"]
    )
    result["totalGG"] = result["planillaStaff"] + result["otrosGG"]

    return result


# ============================================================
# EXTRACT RVAL
# Porta extractRval() del index.html
# ============================================================

def _extract_rval(ws):
    """Extrae datos de la hoja RVAL."""
    result = {
        "projectName": "",
        "date": None,
        "author": "",
        "costoDirecto": 0,
        "gastosGenerales": 0,
        "ggPercent": 0,
        "utilidad": 0,
        "utilPercent": 0,
        "totalValorizacion": 0,
    }

    # Project info
    for r in range(2, 10):
        b = _cell_val(ws, f"B{r}")
        c = _cell_val(ws, f"C{r}")
        g = _cell_val(ws, f"G{r}")

        if b and isinstance(b, str) and "Proyecto" in b:
            result["projectName"] = str(c) if c else ""
        if b and isinstance(b, str) and "Elaborado" in b:
            result["author"] = str(c) if c else ""

        f_val = _cell_val(ws, f"F{r}")
        if f_val and isinstance(f_val, str) and "fecha" in f_val.lower() and g:
            if isinstance(g, datetime):
                result["date"] = g
            elif isinstance(g, (int, float)) and 40000 < g < 60000:
                from datetime import timedelta
                result["date"] = datetime(1899, 12, 30) + timedelta(days=int(g))

        if f_val and isinstance(f_val, str) and "COSTO DIRECTO" in f_val.upper():
            result["costoDirecto"] = _cell_num(ws, f"G{r}")

    # Find actual max row
    actual_max_row = ws.max_row or 103

    # Scan ALL rows from 10 to max for summary labels
    for r in range(10, actual_max_row + 1):
        label_text = None
        for col in ["C", "B", "D"]:
            v = _cell_val(ws, f"{col}{r}")
            if v and isinstance(v, str) and len(v.strip()) > 3:
                label_text = v
                break

        if not label_text:
            continue

        upper = str(label_text).upper().strip()

        # Costo Directo
        if (upper == "COSTO DIRECTO" or
            ("COSTO DIRECTO" in upper and "GASTOS" not in upper
             and "TOTAL" not in upper and len(upper) < 25)):
            for try_col in ["G", "F", "H"]:
                v = _cell_num(ws, f"{try_col}{r}")
                if v > 0:
                    result["costoDirecto"] = v
                    break

        # Gastos Generales
        if "GASTOS GENERALES" in upper:
            for try_col in ["G", "F", "H"]:
                v = _cell_num(ws, f"{try_col}{r}")
                if v > 0:
                    result["gastosGenerales"] = v
                    break
            # Extraer porcentaje del texto
            import re
            match = re.search(r"\(([\d.]+)%?\)", str(label_text))
            if match:
                result["ggPercent"] = float(match.group(1))

        # Utilidad
        if "UTILIDAD" in upper and "TOTAL" not in upper:
            for try_col in ["G", "F", "H"]:
                v = _cell_num(ws, f"{try_col}{r}")
                if v > 0:
                    result["utilidad"] = v
                    break
            import re
            match = re.search(r"\(([\d.]+)%?\)", str(label_text))
            if match:
                result["utilPercent"] = float(match.group(1))

        # Total Valorizacion
        if "TOTAL" in upper and any(kw in upper for kw in ["VALORIZACION", "VALORIZACIÓN", "VALORIZ"]):
            for try_col in ["G", "F", "H"]:
                v = _cell_num(ws, f"{try_col}{r}")
                if v > 0:
                    result["totalValorizacion"] = v
                    break

    # Fallback: derive from percentages if values still 0
    if result["costoDirecto"] > 0:
        if result["gastosGenerales"] == 0 and result["ggPercent"] > 0:
            result["gastosGenerales"] = result["costoDirecto"] * (result["ggPercent"] / 100)
        if result["utilidad"] == 0 and result["utilPercent"] > 0:
            result["utilidad"] = result["costoDirecto"] * (result["utilPercent"] / 100)

    # Calculate percentages if not found
    if not result["ggPercent"] and result["costoDirecto"] > 0:
        result["ggPercent"] = (result["gastosGenerales"] / result["costoDirecto"]) * 100
    if not result["utilPercent"] and result["costoDirecto"] > 0:
        result["utilPercent"] = (result["utilidad"] / result["costoDirecto"]) * 100

    return result


# ============================================================
# EXTRACT CURVA
# Porta extractCurva() del index.html
# ============================================================

def _extract_curva(ws):
    """Extrae datos de la hoja CURVA."""
    result = {
        "contractual": [],
        "valorizado": [],
        "proyectado": None,
        "mesActualIndex": -1,
        "total": 0,
    }

    # Default column layout: PROG=A-E (cols 0-4), EJEC=G-K (cols 6-10)
    prog_start_col = 0  # A=0
    ejec_start_col = 6  # G=6
    plan_start_col = -1

    # Check for PLANIFICADO header in row 1
    for c in range(0, 21):
        addr = f"{_col_letter(c)}1"
        v = _cell_val(ws, addr)
        if v and isinstance(v, str) and "PLANIFICAD" in v.upper():
            plan_start_col = c

    # Read data rows (starting from row 6)
    max_row = ws.max_row or 24

    for r in range(6, max_row + 1):
        prog_mes_addr = f"{_col_letter(prog_start_col)}{r}"
        prog_mes = _cell_val(ws, prog_mes_addr)

        if not prog_mes:
            continue

        prog_mes_str = str(prog_mes).strip()

        if prog_mes_str.upper() == "TOTAL":
            # Total row
            result["total"] = (
                _cell_num(ws, f"{_col_letter(prog_start_col + 1)}{r}")
                or _cell_num(ws, f"{_col_letter(prog_start_col + 2)}{r}")
            )
            continue

        # Contractual (Programado en Excel)
        result["contractual"].append({
            "mes": prog_mes_str,
            "parcial": _cell_num(ws, f"{_col_letter(prog_start_col + 1)}{r}"),
            "acumulado": _cell_num(ws, f"{_col_letter(prog_start_col + 2)}{r}"),
            "parcialPct": _cell_num(ws, f"{_col_letter(prog_start_col + 3)}{r}"),
            "acumPct": _cell_num(ws, f"{_col_letter(prog_start_col + 4)}{r}"),
        })

        # Valorizado (Ejecutado en Excel)
        ejec_parcial = _cell_num(ws, f"{_col_letter(ejec_start_col + 1)}{r}")
        ejec_acum = _cell_num(ws, f"{_col_letter(ejec_start_col + 2)}{r}")
        ejec_parcial_pct = _cell_num(ws, f"{_col_letter(ejec_start_col + 3)}{r}")
        ejec_acum_pct = _cell_num(ws, f"{_col_letter(ejec_start_col + 4)}{r}")

        ejec_mes_val = _cell_val(ws, f"{_col_letter(ejec_start_col)}{r}")
        result["valorizado"].append({
            "mes": str(ejec_mes_val) if ejec_mes_val else prog_mes_str,
            "parcial": ejec_parcial,
            "acumulado": ejec_acum,
            "parcialPct": ejec_parcial_pct,
            "acumPct": ejec_acum_pct,
        })

        # Track last month with valorizado data
        if ejec_parcial > 0 or ejec_acum > 0:
            result["mesActualIndex"] = len(result["contractual"]) - 1

        # Proyectado (Planificado en Excel)
        if plan_start_col >= 0:
            if result["proyectado"] is None:
                result["proyectado"] = []
            plan_mes_val = _cell_val(ws, f"{_col_letter(plan_start_col)}{r}")
            result["proyectado"].append({
                "mes": str(plan_mes_val) if plan_mes_val else prog_mes_str,
                "parcial": _cell_num(ws, f"{_col_letter(plan_start_col + 1)}{r}"),
                "acumulado": _cell_num(ws, f"{_col_letter(plan_start_col + 2)}{r}"),
                "parcialPct": _cell_num(ws, f"{_col_letter(plan_start_col + 3)}{r}"),
                "acumPct": _cell_num(ws, f"{_col_letter(plan_start_col + 4)}{r}"),
            })

    # If total not found from TOTAL row, use last contractual acumulado
    if result["total"] == 0 and result["contractual"]:
        result["total"] = result["contractual"][-1]["acumulado"]

    return result


# ============================================================
# UTILIDADES
# ============================================================

def _get_short_name(full_name):
    """Obtiene nombre corto de la obra. Porta getShortName() del JS."""
    names = [
        "ALMA MATER", "MARA", "CENEPA", "BEETHOVEN",
        "BIOMEDICAS", "BIOMEDIC", "FRANKLIN", "ROOSEVELT",
    ]
    upper = (full_name or "").upper()
    for n in names:
        if n in upper:
            return n
    return full_name[:30] if full_name else "PROYECTO"


def detect_obra_name(excel_bytes):
    """
    Detecta el nombre de la obra del contenido del Excel.
    Util cuando no se puede detectar desde el asunto del correo.

    Returns:
        Nombre del proyecto encontrado, o None
    """
    try:
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)

        # Intentar leer de RES-COSTO
        if "RES-COSTO" in wb.sheetnames:
            ws = wb["RES-COSTO"]
            for r in range(2, 9):
                b = _cell_val(ws, f"B{r}")
                c = _cell_val(ws, f"C{r}")
                if b and isinstance(b, str) and "Proyecto" in b and c:
                    wb.close()
                    return str(c)

        # Intentar leer de RVAL
        if "RVAL" in wb.sheetnames:
            ws = wb["RVAL"]
            for r in range(2, 10):
                b = _cell_val(ws, f"B{r}")
                c = _cell_val(ws, f"C{r}")
                if b and isinstance(b, str) and "Proyecto" in b and c:
                    wb.close()
                    return str(c)

        wb.close()
    except Exception:
        pass

    return None
